from flask import Blueprint, render_template, request, redirect, url_for, session
import openpyxl

products_bp = Blueprint("products", __name__, template_folder="../../templates/products")

FILE = "data_new.xlsx"

def get_sheet(name):
    wb = openpyxl.load_workbook(FILE)
    return wb, wb[name]

@products_bp.route("/", methods=["GET"])
def list_products():
    wb, sheet = get_sheet("Products")
    products = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
    return render_template("products/products.html", products=products, role=session.get("role"))

@products_bp.route("/add", methods=["GET", "POST"])
def add_product():
    if session.get("role") != "Seller":
        return "Unauthorized"
    if request.method == "POST":
        name = request.form["name"]
        price = float(request.form["price"])
        stock = int(request.form["stock"])

        wb, sheet = get_sheet("Products")
        product_id = sheet.max_row
        sheet.append([product_id, name, price, stock, "seller"])
        wb.save(FILE)
        return redirect(url_for("products.list_products"))
    return render_template("products/products.html")

@products_bp.route("/update/<int:pid>", methods=["POST"])
def update_product(pid):
    if session.get("role") != "Seller":
        return "Unauthorized"

    price = float(request.form["price"])
    stock = int(request.form["stock"])

    wb, sheet = get_sheet("Products")
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == pid:
            row[2].value = price
            row[3].value = stock
            break
    wb.save(FILE)
    return redirect(url_for("products.list_products"))
