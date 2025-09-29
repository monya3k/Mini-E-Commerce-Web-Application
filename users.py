from flask import Blueprint, render_template, request, redirect, url_for, session
import openpyxl
from datetime import datetime

users_bp = Blueprint("users", __name__, template_folder="../../templates/users")

FILE = "data_new.xlsx"

def get_sheet(name):
    wb = openpyxl.load_workbook(FILE)
    return wb, wb[name]

@users_bp.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        email = request.form["email"]

        wb, sheet = get_sheet("Users")
        user_id = sheet.max_row
        sheet.append([user_id, username, password, email, "Customer"])
        wb.save(FILE)
        return redirect(url_for("users.login"))
    return render_template("users/register.html")

@users_bp.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]

        # Admin fixed login
        if username == "admin" and password == "admin123":
            session["role"] = "Admin"
            return redirect(url_for("users.dashboard"))

        # Seller fixed login
        if username == "seller" and password == "seller123":
            session["role"] = "Seller"
            return redirect(url_for("users.dashboard"))

        # Customer login
        wb, sheet = get_sheet("Users")
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == username and row[2] == password:
                session["role"] = "Customer"
                session["user_id"] = row[0]
                return redirect(url_for("users.dashboard"))

        return "Invalid credentials!"
    return render_template("users/login.html")

@users_bp.route("/dashboard")
def dashboard():
    if "role" not in session:
        return redirect(url_for("users.login"))
    return render_template("users/dashboard.html", role=session["role"])
