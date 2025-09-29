from flask import Blueprint, request, redirect, url_for, session, render_template
import openpyxl
from datetime import datetime

orders_bp = Blueprint("orders", __name__, template_folder="../../templates/orders")

FILE = "data_new.xlsx"

def get_sheet(name):
    wb = openpyxl.load_workbook(FILE)
    return wb, wb[name]

@orders_bp.route("/place/<int:pid>", methods=["POST"])
def place_order(pid):
    if session.get("role") != "Customer":
        return "Unauthorized"

    qty = int(request.form["quantity"])
    wb, orders = get_sheet("Orders")
    _, products = get_sheet("Products")

    # Update product stock
    for row in products.iter_rows(min_row=2):
        if row[0].value == pid:
            if row[3].value < qty:
                return "Not enough stock!"
            row[3].value -= qty
            break

    order_id = orders.max_row
    orders.append([order_id, session["user_id"], pid, qty, datetime.now().strftime("%Y-%m-%d %H:%M")])

    wb.save(FILE)
    return redirect(url_for("products.list_products"))

@orders_bp.route("/all")
def all_orders():
    if session.get("role") != "Admin":
        return "Unauthorized"

    wb, sheet = get_sheet("Orders")
    orders = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
    return render_template("orders/orders.html", orders=orders)
